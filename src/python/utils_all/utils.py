import json
import os
import sys
from os.path import join as fullfile
import numpy as np
import cv2 as cv
import math
import random
import torch
import torch.nn as nn
from matplotlib import pyplot as plt
from torch.nn import DataParallel
from .import pytorch_ssim
# from pytorch_ssim import *
from torch.utils.data import DataLoader, ConcatDataset
from datetime import datetime
import torch.nn.functional as F
import inspect
from torch.optim.lr_scheduler import StepLR, LambdaLR
from torchvision import transforms as cv_transforms
from openpyxl import Workbook, load_workbook
# from Datasets import SimpleDataset
# from Datasets import SimpleDataset
from . import Datasets
# from FF_CompenUltra_lab.src.python import Datasets
import torchvision.utils as vutils
from torchvision.transforms.functional import to_pil_image
from pathlib import Path
from typing import List, Optional


def make_colorwheel():
    RY, YG, GC, CB, BM, MR = 15, 6, 4, 11, 13, 6
    ncols = RY + YG + GC + CB + BM + MR
    colorwheel = np.zeros((ncols, 3))
    col = 0
    colorwheel[0:RY, 0] = 255
    colorwheel[0:RY, 1] = np.floor(255 * np.arange(0, RY) / RY)
    col += RY
    colorwheel[col:col + YG, 0] = 255 - np.floor(255 * np.arange(0, YG) / YG)
    colorwheel[col:col + YG, 1] = 255
    col += YG
    colorwheel[col:col + GC, 1] = 255
    colorwheel[col:col + GC, 2] = np.floor(255 * np.arange(0, GC) / GC)
    col += GC
    colorwheel[col:col + CB, 1] = 255 - np.floor(255 * np.arange(0, CB) / CB)
    colorwheel[col:col + CB, 2] = 255
    col += CB
    colorwheel[col:col + BM, 2] = 255
    colorwheel[col:col + BM, 0] = np.floor(255 * np.arange(0, BM) / BM)
    col += BM
    colorwheel[col:col + MR, 2] = 255 - np.floor(255 * np.arange(0, MR) / MR)
    colorwheel[col:col + MR, 0] = 255
    return colorwheel


def flow_uv_to_colors(u, v, convert_to_bgr=False):
    flow_image = np.zeros((u.shape[0], u.shape[1], 3), np.uint8)
    colorwheel = make_colorwheel()
    ncols = colorwheel.shape[0]
    rad = np.sqrt(np.square(u) + np.square(v))
    a = np.arctan2(-v, -u) / np.pi
    fk = (a + 1) / 2 * (ncols - 1)
    k0 = np.floor(fk).astype(np.int32)
    k1 = k0 + 1
    k1[k1 == ncols] = 0
    f = fk - k0
    for i in range(colorwheel.shape[1]):
        tmp = colorwheel[:, i]
        col0 = tmp[k0] / 255.0
        col1 = tmp[k1] / 255.0
        col = (1 - f) * col0 + f * col1
        idx = (rad <= 1)
        col[idx] = 1 - rad[idx] * (1 - col[idx])
        col[~idx] = col[~idx] * 0.75
        ch_idx = 2 - i if convert_to_bgr else i
        flow_image[:, :, ch_idx] = np.floor(255 * col)
    return flow_image


def save_flow_to_vismap(flow_tensor, save_path, max_flow=None):
    """
    Enhanced visualization mode:
    - Moderate saturation: controlled around 180 for vivid but readable colors
    - Strong contrast: bright background with clear deformation regions
    """
    # 1. Format conversion [1, 2, H, W] -> [H, W, 2]
    flow_uv = flow_tensor.detach().cpu().numpy()[0].transpose(1, 2, 0)
    u, v = flow_uv[:, :, 0], flow_uv[:, :, 1]

    # 2. Compute polar coordinates
    mag, ang = cv.cartToPolar(u, v)

    # 3. Determine the displacement reference upper bound
    if max_flow is None:
        # Use the 98th percentile to keep the overall tone stable
        reference_max = np.percentile(mag, 98) + 1e-5
    else:
        reference_max = max_flow

    h, w = mag.shape
    hsv = np.zeros((h, w, 3), dtype=np.uint8)

    # [H - Hue] map direction to (0-179)
    hsv[..., 0] = ang * 180 / np.pi / 2

    # [S - Saturation] core adjustment: dynamic range 30-180
    # Linearly map to 0-1
    sat_normalized = np.clip(mag / reference_max, 0, 1)
    # Map to the range from 30 (static) to 180 (maximum displacement)
    # Use power(0.9) to slightly boost the visibility of low-to-mid motion
    sat = (np.power(sat_normalized, 0.9) * 150) + 30
    hsv[..., 1] = sat.astype(np.uint8)

    # [V - Value] keep it at 250 for a bright and clear result
    hsv[..., 2] = 250

    # 5. Convert and save (use BGR for cv2.imwrite compatibility)
    vis_image = cv.cvtColor(hsv, cv.COLOR_HSV2BGR)

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    cv.imwrite(save_path, vis_image)

def get_linux_style_dataset_list(
    dataset_root: str,
    exclude_prefixes: Optional[List[str]] = None,
    include_prefixes: Optional[List[str]] = None
) -> List[str]:
    """
    Return all paths under the `setups` subdirectory that satisfy the given conditions,
    formatted in Linux style (using '/' as the separator).

    Args:
        dataset_root (str): Root directory path, for example 'H:/Valid_datasets/CompenHR_datasets'
        exclude_prefixes (List[str], optional): Exclude directory names with these prefixes
        include_prefixes (List[str], optional): Only include directory names with these prefixes (higher priority than exclude)

    Returns:
        List[str]: A list of paths such as ['setups/Block1', 'setups/Fruits_Vegetables2', ...]
    """
    exclude_prefixes = exclude_prefixes or []
    setups_dir = Path(dataset_root) / 'setups'

    if not setups_dir.exists():
        print(f"[Warning] Directory does not exist: {setups_dir}")
        return []

    all_dirs = [name for name in os.listdir(setups_dir) if (setups_dir / name).is_dir()]

    if include_prefixes:
        filtered_dirs = [
            name for name in all_dirs if any(name.startswith(p) for p in include_prefixes)
        ]
    else:
        filtered_dirs = [
            name for name in all_dirs if not any(name.startswith(p) for p in exclude_prefixes)
        ]

    data_list = [os.path.join("setups", name).replace("\\", "/") for name in filtered_dirs]

    # ✅ Print the output list
    print("Dataset list:")
    for path in data_list:
        print(f"  {path}")
    print(f"\nTotal: {len(data_list)} setup directories")

    return data_list

def visualize_warp(original, warped, mask, warped_clean, tag='debug', save_path='./debug'):
    os.makedirs(save_path, exist_ok=True)
    b = original.size(0)
    for i in range(b):
        orig_img = to_pil_image(original[i].cpu())
        warped_img = to_pil_image(warped[i].cpu())
        mask_img = to_pil_image(mask[i].expand(3, -1, -1).cpu())  # Expand a single-channel mask to three channels
        warped_clean_img = to_pil_image(warped_clean[i].cpu())

        orig_img.save(os.path.join(save_path, f'{tag}_b{i}_orig.png'))
        warped_img.save(os.path.join(save_path, f'{tag}_b{i}_warped.png'))
        mask_img.save(os.path.join(save_path, f'{tag}_b{i}_mask.png'))
        warped_clean_img.save(os.path.join(save_path, f'{tag}_b{i}_warpedclean.png'))

def load_model_with_bias_resize(model, checkpoint_path):
    checkpoint = torch.load(checkpoint_path)
    state_dict = checkpoint if "state_dict" not in checkpoint else checkpoint["state_dict"]

    for k in list(state_dict.keys()):
        if "relative_position_bias_table" in k:
            pretrained_bias = state_dict[k]
            current_bias = model.state_dict()[k]
            if pretrained_bias.shape != current_bias.shape:
                print(f"↔️ Auto interpolation: {k} from {pretrained_bias.shape} to {current_bias.shape}")
                # The old shape is [L, num_heads]
                num_heads = current_bias.shape[1]
                old_len = int(pretrained_bias.shape[0] ** 0.5)
                new_len = int(current_bias.shape[0] ** 0.5)
                pretrained_bias = pretrained_bias.reshape(1, num_heads, old_len, old_len)
                interpolated = torch.nn.functional.interpolate(
                    pretrained_bias,
                    size=(new_len, new_len),
                    mode="bicubic",
                    align_corners=False,
                )
                state_dict[k] = interpolated.reshape(current_bias.shape)

    model.load_state_dict(state_dict, strict=False)

data_transforms = {'surf': None,  # surf does not need conversion because it has already been processed into the PyTorch training format in the dataset
                   'cam': cv_transforms.Compose([cv_transforms.ToTensor()]),
                   'prj': cv_transforms.Compose([cv_transforms.ToTensor()])}


def worker_init_fn(worker_id, cfg):
    worker_seed = cfg.trainer.randseed + worker_id
    np.random.seed(worker_seed)
    random.seed(worker_seed)
    torch.manual_seed(worker_seed)

# def create_dataloader(train_datasets, test_dataset, cfg):
#     # Create the data generator
#     data_gen = torch.Generator()
#     data_gen.manual_seed(cfg.trainer.randseed)
#
#     # --- Key modification: use partial to wrap the function ---
#     # This creates a new function that automatically passes cfg to worker_init_fn
#     wrapped_init_fn = functools.partial(worker_init_fn, cfg=cfg)
#
#     # Merge training datasets
#     Concat_dataset = ConcatDataset(train_datasets)
#     Concat_dataloader = DataLoader(
#         Concat_dataset,
#         batch_size=cfg.batch_size,
#         shuffle=True,
#         num_workers=cfg.num_workers,
#         generator=data_gen,
#         worker_init_fn=wrapped_init_fn  # Use the wrapped function
#     )
#
#     # Create the validation dataset
#     test_loader = DataLoader(
#         test_dataset,
#         batch_size=cfg.batch_size,
#         shuffle=True,
#         num_workers=cfg.num_workers,
#         generator=data_gen,
#         worker_init_fn=wrapped_init_fn  # Use the wrapped function
#     )
#
#     return Concat_dataloader, test_loader
# Linux does not require serialization here
def create_dataloader(train_datasets, test_dataset, cfg):
    # Create the data generator
    data_gen = torch.Generator()
    data_gen.manual_seed(cfg.trainer.randseed)

    # Merge training datasets
    Concat_dataset = ConcatDataset(train_datasets)
    Concat_dataloader = DataLoader(
        Concat_dataset,
        batch_size=cfg.batch_size,
        shuffle=True,
        num_workers=cfg.num_workers,
        generator=data_gen,
        worker_init_fn=lambda worker_id: worker_init_fn(worker_id, cfg)  # Pass cfg to worker_init_fn
    )

    # Create the validation dataset
    test_loader = DataLoader(
        test_dataset,
        batch_size=cfg.batch_size,
        shuffle=True,
        num_workers=cfg.num_workers,
        generator=data_gen,
        worker_init_fn=lambda worker_id: worker_init_fn(worker_id, cfg)  # Pass cfg to worker_init_fn as well
    )

    return Concat_dataloader, test_loader


def getLargestRect(imBW, aspectRatio):
    h, w = imBW.shape

    # find candidate rect centers using distance transform
    imDist = cv.distanceTransform(imBW, cv.DIST_L2, 3)
    cy, cx = np.where(imDist > np.percentile(imDist[imBW == 0], 90))

    # only use a subset for speedup
    np.random.seed(1)
    idx = np.random.choice(len(cx), min(200, len(cx)), replace=False)
    cx = cx[idx]
    cy = cy[idx]

    maxH = 0

    for i in range(len(cy)):
        lx, rx, ty, by = cx[i], cx[i], cy[i], cy[i]

        while True:
            curH = by - ty + 1
            dx = round(curH * aspectRatio) - round((curH - 1) * aspectRatio)

            if ty - 1 >= 0 and lx - dx >= 0:
                imRect = imBW[ty - 1:by + 1, lx - dx:rx + 1]
                if np.count_nonzero(imRect == 0) == 0:
                    ty -= 1
                    lx -= dx
                else:
                    break
            else:
                break

            curH = by - ty + 1
            dx = round(curH * aspectRatio) - round((curH - 1) * aspectRatio)

            if by + 1 < h and rx + dx < w:
                imRect = imBW[ty:by + 2, lx:rx + dx + 2]
                if np.count_nonzero(imRect == 0) == 0:
                    by += 1
                    rx += dx
                else:
                    break
            else:
                break

        curH = by - ty + 1
        if curH > maxH:
            maxH = curH
            bestx, besty = cx[i], cy[i]
            bestRect = [lx, ty, rx - lx + 1, by - ty + 1]
        elif curH == maxH:
            bestx = np.append(bestx, cx[i])
            besty = np.append(besty, cy[i])
            bestRect = np.vstack((bestRect, [lx, ty, rx - lx + 1, by - ty + 1]))

    idx = np.argmax(imDist[besty, bestx])
    rect = bestRect[idx]

    return rect


# set random number generators' seeds
def resetRNGseed(seed):
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def Get_dataLoader(img_dir, size=None, index=None, cfg=None):
    img_dataset = Datasets.SimpleDataset(img_dir, index=index, size=size)
    data_loader = DataLoader(img_dataset, batch_size=cfg.batch_size, shuffle=False, drop_last=False, num_workers=0)
    return data_loader


# Same as np.repeat, while torch.repeat works as np.tile
def repeat_np(a, repeats, dim):
    '''
    Substitute for numpy's repeat function. Source from https://discuss.pytorch.org/t/how-to-tile-a-tensor/13853/2
    torch.repeat([1,2,3], 2) --> [1, 2, 3, 1, 2, 3]
    np.repeat([1,2,3], repeats=2, axis=0) --> [1, 1, 2, 2, 3, 3]

    :param a: tensor
    :param repeats: number of repeats
    :param dim: dimension where to repeat
    :return: tensor with repitions
    '''

    init_dim = a.size(dim)
    repeat_idx = [1] * a.dim()
    repeat_idx[dim] = repeats
    a = a.repeat(*(repeat_idx))
    if a.is_cuda:  # use cuda-device if input was on cuda device already
        order_index = torch.cuda.LongTensor(
            torch.cat([init_dim * torch.arange(repeats, device=a.device) + i for i in range(init_dim)]))
    else:
        order_index = torch.LongTensor(
            torch.cat([init_dim * torch.arange(repeats) + i for i in range(init_dim)]))

    return torch.index_select(a, dim, order_index)


# save 4D np.ndarray or torch tensor to image files
def saveImgs(inputData, dir):
    if not os.path.exists(dir):
        os.makedirs(dir)

    if type(inputData) is torch.Tensor:
        if inputData.requires_grad:
            inputData = inputData.detach()
        if inputData.device.type == 'cuda':
            imgs = inputData.cpu().numpy().transpose(0, 2, 3, 1)  # (N, C, row, col) to (N, row, col, C)
        else:
            imgs = inputData.numpy().transpose(0, 2, 3, 1)  # (N, C, row, col) to (N, row, col, C)

    else:
        imgs = inputData

    # imgs must have a shape of (N, row, col, C)
    imgs = np.uint8(imgs[:, :, :, ::-1] * 255)  # convert to BGR and uint8 for opencv
    for i in range(imgs.shape[0]):
        file_name = 'img_{:04d}.png'.format(i + 1)
        cv.imwrite(fullfile(dir, file_name), imgs[i, :, :, :])  # faster than PIL or scipy


# compute PSNR
def psnr(x, y):
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    x = x.to(device) if x.device.type != device.type else x
    y = y.to(device) if y.device.type != device.type else y

    with torch.no_grad():
        l2_fun = nn.MSELoss()
        return 10 * math.log10(1 / l2_fun(x, y))


# compute RMSE
def rmse(x, y):
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    x = x.to(device) if x.device.type != device.type else x
    y = y.to(device) if y.device.type != device.type else y

    with torch.no_grad():
        l2_fun = nn.MSELoss()
        return math.sqrt(l2_fun(x, y).item() * 3)  # only works for RGB, for grayscale, don't multiply by 3


# compute ssim
def ssim(x, y):
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    x = x.to(device) if x.device.type != device.type else x
    y = y.to(device) if y.device.type != device.type else y

    with torch.no_grad():
        return pytorch_ssim.ssim(x, y).item()


# compute psnr, rmse and ssim
def computeMetrics(x, y) -> object:
    l2_fun = nn.MSELoss()
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    last_loc = 0
    metric_mse, metric_ssim = 0., 0.

    num_imgs = x.shape[0]
    batch_size = 50 if num_imgs > 50 else num_imgs  # Make sure mod(num_imgs, batch_size) == 0

    with torch.no_grad():
        for i in range(0, num_imgs // batch_size):
            idx = range(last_loc, last_loc + batch_size)
            x_batch = x[idx, :, :, :].to(device) if x.device.type != 'cuda' else x[idx, :, :, :]
            y_batch = y[idx, :, :, :].to(device) if y.device.type != 'cuda' else y[idx, :, :, :]

            # compute mse and ssim
            metric_mse += l2_fun(x_batch, y_batch).item() * batch_size
            metric_ssim += ssim(x_batch, y_batch) * batch_size

            last_loc += batch_size

        # average
        metric_mse /= num_imgs
        metric_ssim /= num_imgs

        # rmse and psnr
        metric_rmse = math.sqrt(metric_mse * 3)  # 3 channel image
        metric_psnr = 10 * math.log10(1 / metric_mse)

    return metric_psnr, metric_rmse, metric_ssim


# count the number of parameters of a model
def countParameters(model):
    return sum([param.numel() for param in model.parameters() if param.requires_grad])


# generate training title string
def optionToString(train_option):
    return '{}_{}_{}_{}_{}_{}_{}_{}_{}_{}'.format(train_option['data_name'], train_option['model_name'],
                                                  train_option['loss'],
                                                  train_option['num_train'], train_option['batch_size'],
                                                  train_option['max_iters'],
                                                  train_option['lr'], train_option['lr_drop_ratio'],
                                                  train_option['lr_drop_rate'],
                                                  train_option['l2_reg'])


def log_surrogate(log_dir, current_time, data_name, model_name, loss_function, num_train,
                  batch_size, valid_psnr, valid_rmse, valid_ssim, valid_diff, valid_lpips):
    """
    Record training results to a log file (the title is only written when the file is first created)

    Args:
        log_dir:        Log directory path
        current_time:   Externally passed time string
        ... (other arguments are the same as before)
    """
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    log_path = os.path.join(log_dir, f"{current_time}_surrogate.txt")

    # Check if the file already exists (decide whether to write the title)
    write_header = not os.path.exists(log_path)

    with open(log_path, 'a') as log_file:
        if write_header:
            title_str = '{:30s}{:<30}{:<20}{:<15}{:<15}{:<15}{:<15}{:<15}{:<15}{:<15}\n'
            log_file.write(title_str.format(
                'data_name', 'model_name', 'loss_function',
                'num_train', 'batch_size',
                'valid_psnr', 'valid_rmse', 'valid_ssim', "valid_diff", "valid_lpips"
            ))

        ret_str = '{:30s}{:<30}{:<20}{:<15}{:<15}{:<15.4f}{:<15.4f}{:<15.4f}{:<15.4f}{:<15.4f}\n'
        log_file.write(ret_str.format(
            data_name, model_name, loss_function,
            str(num_train), str(batch_size),
            valid_psnr, valid_rmse, valid_ssim, valid_diff, valid_lpips
        ))

    print(f"Results appended to: {log_path}")


def save_test_log_one_line(folder_path, model_name, batch_size, lr, loss, uncmp_psnr, uncmp_rmse,
                           uncmp_ssim, uncmp_deltaE, uncmp_lpips, psnr, rmse, ssim, deltaE,
                           lpips, txt_file_name="test_log.txt", excel_file_name="test_log.xlsx"):
    """
    Save training logs to specified folder path in both text file and Excel file formats, appending new rows on each call.
    Learning rate (lr) is retained with 7 decimal places.

    :param folder_path: Log file folder path.
    :param model_name: Model name.
    :param batch_size: Batch size.
    :param lr: Learning rate.
    :param loss: Current loss description string.
    :param uncmp_psnr: Original PSNR metric.
    :param uncmp_rmse: Original RMSE metric.
    :param uncmp_ssim: Original SSIM metric.
    :param uncmp_deltaE: Original DeltaE metric.
    :param uncmp_lpips: Original LPIPS metric.
    :param psnr: Current PSNR value.
    :param rmse: Current RMSE value.
    :param ssim: Current SSIM value.
    :param deltaE: Current DeltaE value.
    :param lpips: Current LPIPS value.
    :param txt_file_name: Text log file name (default is "test_log.txt").
    :param excel_file_name: Excel log file name (default is "test_log.xlsx").
    """
    try:
        # Full path for the log files
        txt_log_path = os.path.join(folder_path, txt_file_name)
        excel_log_path = os.path.join(folder_path, excel_file_name)

        # Ensure the folder exists (create recursively if not)
        os.makedirs(folder_path, exist_ok=True)

        # Define the title and width for each column
        title = [
            "Time", "Model Name", "Batch_Size", "Learning_Rate", "Loss",
            "UnCmp_PSNR", "UnCmp_RMSE", "UnCmp_SSIM",
            "UnCmp_DeltaE", "UnCmp_LPIPS",
            "PSNR", "RMSE", "SSIM", "DeltaE", "LPIPS"
        ]
        widths = [20, 25, 15, 20, 20, 20, 20, 20, 20, 20, 12, 12, 12, 12, 12]

        # ===== Save to text file =====
        # Check if the text file exists
        if not os.path.exists(txt_log_path):
            with open(txt_log_path, "w", encoding="utf-8") as txt_file:
                # Format the title to fixed column widths and center align
                title_line = "".join([col.center(width) for col, width in zip(title, widths)]) + "\n"
                txt_file.write(title_line)
                # print(f"Log TXT file created with headers: {txt_log_path}")

        # Get the current time string
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Format the log content to fixed column widths and center align, learning rate with 7 decimal places
        log_message = "".join([
            str(current_time).center(widths[0]),
            str(model_name).center(widths[1]),
            str(batch_size).center(widths[2]),
            f"{lr:.7f}".center(widths[3]),  # Modified: 7 decimal places
            str(loss).center(widths[4]),
            f"{uncmp_psnr:.4f}".center(widths[5]),
            f"{uncmp_rmse:.4f}".center(widths[6]),
            f"{uncmp_ssim:.4f}".center(widths[7]),
            f"{uncmp_deltaE:.4f}".center(widths[8]),
            f"{uncmp_lpips:.4f}".center(widths[9]),
            f"{psnr:.4f}".center(widths[10]),
            f"{rmse:.4f}".center(widths[11]),
            f"{ssim:.4f}".center(widths[12]),
            f"{deltaE:.4f}".center(widths[13]),
            f"{lpips:.4f}".center(widths[14]) + "\n"
        ])

        # Append the log content to the text file
        with open(txt_log_path, "a", encoding="utf-8") as txt_file:
            txt_file.write(log_message)

        # print(f"Log entry appended to TXT file: {txt_log_path}")

        # ===== Save to Excel file =====
        # Check if the Excel file exists
        if not os.path.exists(excel_log_path):
            # Create a new Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Test_Log"  # Optional: set the sheet name
            ws.append(title)  # Write the title
            wb.save(excel_log_path)
            # print(f"Log Excel file created with headers: {excel_log_path}")

        # Open the Workbook
        wb = load_workbook(excel_log_path)
        ws = wb.active

        # Create the log data row, learning rate with 7 decimal places
        log_row = [
            current_time,
            model_name,
            batch_size,
            round(lr, 7),  # Modified: 7 decimal places
            loss,
            round(uncmp_psnr, 4),
            round(uncmp_rmse, 4),
            round(uncmp_ssim, 4),
            round(uncmp_deltaE, 4),
            round(uncmp_lpips, 4),
            round(psnr, 4),
            round(rmse, 4),
            round(ssim, 4),
            round(deltaE, 4),
            round(lpips, 4)
        ]

        # Append the new row
        ws.append(log_row)

        # Save the Workbook
        wb.save(excel_log_path)

        # print(f"Log entry appended to Excel file: {excel_log_path}")

    except Exception as e:
        print(f"Error saving log files: {e}")


def save_vis_line(vis, win, save_path):
    """
    Save Visdom chart as an image, displaying only the last data point of the curve, and not connecting the 0th round point to the origin.

    :param vis: Visdom instance
    :param win: Visdom chart window name
    :param save_path: Path to save the image
    """
    # Get the data from the window
    window_data = vis.get_window_data(win)
    if not window_data:  # Check if the data is empty
        print(f"Warning: No data found in Visdom window '{win}'")
        return

    try:
        # Try to parse the window data
        window_json = json.loads(window_data)
        content = window_json.get('content', {})
        traces = content.get('data', [])  # The curve data in the chart is usually in 'data'

        # Create the plot
        plt.figure(figsize=(10, 6))
        for trace in traces:
            x = trace['x']  # x-axis data
            y = trace['y']  # y-axis data
            name = trace.get('name', 'unknown')  # Get the curve name

            # Plot only if there is more than 1 data point
            if len(x) > 0 and len(y) > 0:
                plt.plot(x[1:], y[1:], label=name)  # Plot from the 1st point
                # plt.scatter(x[0:1], y[0:1], color='red')  # Plot the 0th round point separately

                # Display the value of the last point
                plt.annotate(
                    f'{y[-1]:.4f}',  # Show only the value of the last point
                    (x[-1], y[-1]),  # Position of the last point
                    textcoords="offset points",  # Offset mode
                    xytext=(0, 5),  # Offset (x and y direction)
                    ha='center',  # Horizontal alignment
                    fontsize=8,  # Font size
                    color='black'  # Font color
                )
        # Get the title
        title = content.get('layout', {}).get('title', {})
        if isinstance(title, dict) and 'text' in title:
            plt.title(title['text'])  # If in dictionary format
        else:
            plt.title(title)  # If in string format

        # Get the x and y axis labels
        xaxis = content.get('layout', {}).get('xaxis', {}).get('title', {})
        plt.xlabel(xaxis['text'] if isinstance(xaxis, dict) and 'text' in xaxis else xaxis)

        yaxis = content.get('layout', {}).get('yaxis', {}).get('title', {})
        plt.ylabel(yaxis['text'] if isinstance(yaxis, dict) and 'text' in yaxis else yaxis)

        # Add legend
        plt.legend()

        # Save the chart as an image
        plt.savefig(save_path)
        plt.close()

    except Exception as e:
        print(f"Error while saving Visdom graph: {e}")


def create_folder_with_time(base_folder, current_time):
    """
    Create two levels of folders under the specified base_folder based on the passed time string:
    - Level 1: Date folder (format 'YYYY_MM_DD')
    - Level 2: Time folder (format 'HH_MM_SS')

    :param base_folder: Top-level folder name (e.g., 'log' or 'SCNet_surf1_img')
    :param current_time: Time string, format 'YYYY_MM_DD_HH_MM_SS'
    :return: The final created time folder path
    """
    try:
        # Check the format of the time string
        if "_" not in current_time:
            raise ValueError(f"Invalid time format: {current_time}. Expected format is 'YYYY_MM_DD_HH_MM_SS'.")

        # Split the time string into date and time parts
        parts = current_time.split("_")
        if len(parts) != 6:
            raise ValueError(f"Invalid time format: {current_time}. Expected format is 'YYYY_MM_DD_HH_MM_SS'.")

        date_part = "_".join(parts[:3])  # Date part (YYYY_MM_DD)
        time_part = "_".join(parts[3:])  # Time part (HH_MM_SS)

        # Build the date folder path
        date_folder_path = os.path.join(base_folder, date_part)

        # Create the date folder if it does not exist
        if not os.path.exists(date_folder_path):
            os.makedirs(date_folder_path)
            print(f"Created date folder: {date_folder_path}")

        # Build the time folder path
        time_folder_path = os.path.join(date_folder_path, time_part)

        # Create the time folder if it does not exist
        if not os.path.exists(time_folder_path):
            os.makedirs(time_folder_path)
            print(f"Created time folder: {time_folder_path}")

        # Return the final created time folder path
        return time_folder_path

    except Exception as e:
        print(f"Error creating folder: {e}")
        return None


def preprocess_4d_tensor(input_tensor, device, repeat_times=None):
    """
    Process a 4D tensor: normalize, adjust channel order, and transfer to the specified device.

    Args:
        input_tensor: The input 4D tensor with shape (N, H, W, C).
        device: The device to transfer the data to (e.g., 'cuda' or 'cpu').
        repeat_times: Repeat times (optional), for repeating tensor dimensions.

    Returns:
        The processed tensor on the specified device.
    """
    # If repetition is needed
    if repeat_times is not None:
        input_tensor = input_tensor.repeat(repeat_times, 1, 1, 1)

    # Normalize and change data type
    processed_tensor = input_tensor.permute(0, 3, 1, 2).float().div(255)

    # Transfer data to the specified device
    processed_tensor = processed_tensor.to(device)

    return processed_tensor


def preprocess_cam_surf_5_data(data_loader, device):
    # Get one batch of data
    data = next(iter(data_loader))

    # Convert to target device and adjust data type
    data = data.to(device=device, dtype=torch.float32)

    # Adjust dimensions and normalize
    data = data.permute(0, 3, 1, 2)
    data = data / 255.0

    # Use reshape to adjust data shape
    data = data.reshape(1, 15, 256, 256)

    return data


def load_state_dict_without_module(pth_file, model):
    """
    Load a state dict from a .pth file, remove 'module.' prefix from the keys,
    and load it into the given model.
    Parameters:
    pth_file (str): Path to the input .pth file.
    model (torch.nn.Module): The model to load the state dict into.
    """
    # Load the state dict
    state_dict = torch.load(pth_file)
    # Remove 'module.' prefix from the keys
    fixed_state_dict = {k.replace("module.", ""): v for k, v in state_dict.items()}
    # Load the fixed state dict into the model
    model.load_state_dict(fixed_state_dict, strict=False)
    print(f"{pth_file} State dict loaded into model successfully")


def flow_saveImgs(inputData, dir, start_idx=0):
    if not os.path.exists(dir):
        os.makedirs(dir)
    if isinstance(inputData, torch.Tensor):
        if inputData.requires_grad:
            inputData = inputData.detach()
        if inputData.device.type == 'cuda':
            imgs = inputData.cpu().numpy().transpose(0, 2, 3, 1)  # (N, C, row, col) to (N, row, col, C)
        else:
            imgs = inputData.numpy().transpose(0, 2, 3, 1)  # (N, C, row, col) to (N, row, col, C)
    else:
        imgs = inputData
    # imgs must have a shape of (N, row, col, C)
    imgs = np.uint8(imgs * 255)  # assume images are normalized to [0, 1]
    for i in range(imgs.shape[0]):
        file_name = 'img_{:04d}.png'.format(start_idx + i + 1)
        cv.imwrite(fullfile(dir, file_name), cv.cvtColor(imgs[i, :, :, :], cv.COLOR_BGR2RGB))


def warp_images(image, flow, mode='bilinear', padding_mode='zeros'):
    """
    Returns the warped image filtered by the mask, ensuring the invalid areas are 0
    Args:
        image (torch.Tensor): (B, C, H, W)
        flow (torch.Tensor): (B, 2, H, W)
        mode (str): 'bilinear' or 'bicubic'
        padding_mode (str): 'zeros', 'border', 'reflection'
    Returns:
        warped_clean (torch.Tensor): (B, C, H, W)
    """
    B, C, H, W = image.size()
    device = image.device
    # Generate standard grid
    grid_x, grid_y = torch.meshgrid(
        torch.linspace(-1.0, 1.0, W, device=device),
        torch.linspace(-1.0, 1.0, H, device=device),
        indexing='xy'
    )
    grid = torch.stack((grid_x, grid_y), dim=-1)  # (H, W, 2)
    grid = grid.unsqueeze(0).repeat(B, 1, 1, 1)  # (B, H, W, 2)
    # Normalize flow
    flow_norm = torch.zeros_like(flow)
    flow_norm[:, 0, :, :] = flow[:, 0, :, :] / ((W - 1) / 2)
    flow_norm[:, 1, :, :] = flow[:, 1, :, :] / ((H - 1) / 2)
    flow_norm = flow_norm.permute(0, 2, 3, 1)
    vgrid = grid + flow_norm
    # grid_sample
    warped = F.grid_sample(image, vgrid, mode=mode, padding_mode=padding_mode, align_corners=True)
    return warped

def save_config_and_model_info(config, model, output_dir):
    """
    Save the configuration file and model-related information to the log file in the specified folder.

    Args:
        config (CN): Configuration object (e.g., _CN).
        model (object): Declared external model object.
        output_dir (str): Folder path to save the log file.
    """
    # Ensure the output directory exists, create if not
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")

    # Log file name
    output_file_path = os.path.join(output_dir, "config_and_model_info.txt")

    # Get the class definition source of the model
    cls = model.__class__
    class_source_code = inspect.getsource(cls)

    # Write configuration and model information to the same file
    with open(output_file_path, 'w') as output_file:
        # Write configuration content
        output_file.write("### Configuration ###\n")
        output_file.write(config.dump())  # Dump the configuration content in YAML format
        output_file.write("\n\n")

        # Write model information
        output_file.write("### Model Information ###\n")
        output_file.write(f"Model Name: {model.__class__.__name__}\n\n")
        output_file.write("### Model Class Definition ###\n")
        output_file.write(class_source_code)

    print(f"Configuration and model information saved to: {output_file_path}")


def get_scheduler(optimizer, scheduler_config):
    """
    Return the corresponding learning rate scheduler based on the configuration
    """
    scheduler_type = scheduler_config.type
    if scheduler_type == 'step_lr':
        return StepLR(optimizer, step_size=scheduler_config.step_size, gamma=scheduler_config.gamma)
    elif scheduler_type == 'lambda_lr':
        a = scheduler_config.a
        lambda_func = lambda it: 1 / (1 + a * it)
        return LambdaLR(optimizer, lr_lambda=lambda_func)
    elif scheduler_type == 'cosine_annealing':
        return torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=scheduler_config.T_max, eta_min=scheduler_config.eta_min)
    else:
        raise ValueError(f"Unsupported scheduler type: {scheduler_type}")


def save_two_checkpoint(model_a, model_b, optimizer_a, optimizer_b, scheduler_a, scheduler_b, total_steps, save_interval, save_dir,
                        is_final=False):
    """
    Save model checkpoints

    Args:
    - model_a: Model A to be saved
    - model_b: Model B to be saved
    - optimizer_a: Optimizer for Model A
    - optimizer_b: Optimizer for Model B
    - scheduler_a: Learning rate scheduler for Model A
    - scheduler_b: Learning rate scheduler for Model B
    - total_steps: Current iteration count
    - save_interval: Save every how many rounds
    - save_dir: Base directory for saving
    - is_final: Whether it is the final save (boolean)
    """

    # Check if checkpoint needs to be saved
    should_save = (total_steps % save_interval == 0) or is_final

    if should_save:
        # Decide file name based on final save or not
        if is_final:
            filename = f'step_{total_steps}_final.pth'
        else:
            filename = f'step_{total_steps}.pth'

        # Define the file path for saving
        final_path = os.path.join(save_dir, filename)

        try:
            # Create a dictionary to save the states of two models, optimizers, and schedulers
            checkpoint = {
                'model_a_state_dict': model_a.module.state_dict() if isinstance(model_a, DataParallel) else model_a.state_dict(),
                'model_b_state_dict': model_b.module.state_dict() if isinstance(model_b, DataParallel) else model_b.state_dict(),
                'optimizer_a_state_dict': optimizer_a.state_dict(),
                'optimizer_b_state_dict': optimizer_b.state_dict(),
                'scheduler_a_state_dict': scheduler_a.state_dict(),
                'scheduler_b_state_dict': scheduler_b.state_dict(),
                'total_steps': total_steps,
            }

            # Save the checkpoint
            torch.save(checkpoint, final_path)
            print(f"Saved checkpoint to: {final_path}")
        except Exception as e:
            print(f"Error saving checkpoint: {e}")


def load_checkpoint(filename, model_a, model_b, optimizer_a, optimizer_b, scheduler_a, scheduler_b):
    """
    Load checkpoints

    Args:
    - filename: Path to the saved checkpoint file
    - model_a: Model A to load
    - model_b: Model B to load
    - optimizer_a: Optimizer for Model A
    - optimizer_b: Optimizer for Model B
    - scheduler_a: Learning rate scheduler for Model A
    - scheduler_b: Learning rate scheduler for Model B

    Returns:
    - total_steps: Loaded total iteration count
    """
    try:
        # Load the checkpoint
        checkpoint = torch.load(filename)

        # Load model state dicts
        model_a.load_state_dict(checkpoint['model_a_state_dict'])
        model_b.load_state_dict(checkpoint['model_b_state_dict'])

        # Load optimizer state dicts
        optimizer_a.load_state_dict(checkpoint['optimizer_a_state_dict'])
        optimizer_b.load_state_dict(checkpoint['optimizer_b_state_dict'])

        # Load scheduler state dicts
        scheduler_a.load_state_dict(checkpoint['scheduler_a_state_dict'])
        scheduler_b.load_state_dict(checkpoint['scheduler_b_state_dict'])

        # Return the loaded total iteration count
        total_steps = checkpoint['total_steps']
        print(f"Successfully loaded checkpoint: {filename} (Total steps: {total_steps})")
        return total_steps

    except Exception as e:
        print(f"Error loading checkpoint: {e}")
        raise


def save_one_model_checkpoint(model, total_steps, save_interval, save_dir, current_time, is_final=False):
    """
    Save model checkpoints

    Args:
    - model: Model to be saved
    - total_steps: Current iteration count
    - save_interval: Save every how many rounds
    - save_dir: Base directory for saving
    - current_time: Current time string for creating independent folders
    - is_final: Whether it is the final save (boolean)
    """

    # Check if checkpoint needs to be saved
    should_save = (total_steps % save_interval == 0) or is_final

    if should_save:
        # Create new save directory (based on current time)
        checkpoint_dir = os.path.join(save_dir, current_time)
        os.makedirs(checkpoint_dir, exist_ok=True)

        # Decide file name based on final save or not
        if is_final:
            filename = f'step_{total_steps}_final.pth'
        else:
            filename = f'step_{total_steps}.pth'

        # Define the file path for saving
        final_path = os.path.join(checkpoint_dir, filename)

        try:
            # If the model uses DataParallel, save the state_dict of the module
            if isinstance(model, DataParallel):
                state_dict = model.module.state_dict()
            else:
                state_dict = model.state_dict()

            # Save the model's state dict
            torch.save(state_dict, final_path)
            print(f"Saved checkpoint to: {final_path}")
        except Exception as e:
            print(f"Error saving checkpoint: {e}")


def import_datasets_module(choice):
    if choice.lower() == 'lab2':
        try:
            import configs.lab2_datasets_lists as datasets_module
            print("Successfully imported module configs.lab2_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.lab2_datasets_lists: {e}")
            sys.exit(1)

    elif choice.lower() == 'sicomp':
        try:
            import python.configs.SIComp_datasets_lists as datasets_module
            print("Successfully imported module SIComp_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.lab3_datasets_lists: {e}")
            sys.exit(1)
    elif choice.lower() == 'local':
        try:
            import python.configs.IVPCNet_datasets_lists as datasets_module
            print("Successfully imported module configs.local_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.local_datasets_lists: {e}")
            sys.exit(1)
    elif choice.lower() == 'sl_lab3':
        try:
            import configs.SL_lab3_datasets_lists as datasets_module
            print("Successfully imported module configs.SL_lab3_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.SL_lab3_datasets_lists: {e}")
            sys.exit(1)

    elif choice.lower() == 'sl_lab2':
        try:
            import configs.SL_lab2_datasets_lists as datasets_module
            print("Successfully imported module configs.SL_lab2_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.SL_lab2_datasets_lists: {e}")
            sys.exit(1)

    elif choice.lower() == 'sl_local':
        try:
            import python.configs.IVPCNet_datasets_lists as datasets_module
            print("Successfully imported module configs.SL_local_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.SL_local_datasets_lists: {e}")
            sys.exit(1)

    elif choice.lower() == 'sl_without_cmp_lab3':
        try:
            import configs.SL_lab3_datasets_lists_without_cmp as datasets_module
            print("Successfully imported module configs.SL_local_datasets_lists")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.SL_local_datasets_lists: {e}")
            sys.exit(1)

    elif choice.lower() == 'sl_without_cmp_local':
        try:
            import configs.SL_local_datasets_lists_without_cmp as datasets_module
            print("Successfully imported module configs.SL_local_datasets_lists_without_cmp")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.SL_local_datasets_lists_without_cmp: {e}")
            sys.exit(1)

    elif choice.lower() == 'test_local':
        try:
            import configs.Test_datasets as datasets_module
            print("Successfully imported module configs.Test_local")
            return datasets_module
        except ImportError as e:
            print(f"Failed to import module configs.Test_local: {e}")
            sys.exit(1)
    else:
        print(f"Unknown dataset_type: '{choice}'. Please use 'lab2' or 'lab3'.")
        sys.exit(1)


def visdom_display_save(vis, cam_crop, warped_predict, prj_GT, batch_size, step, phase='train',
                        max_nrow=5, save_path=None, win=None):
    """
    Visualize the stitched images and display them in the Visdom window, with an option to save the images locally.

    Args:
        vis (visdom.Visdom): Initialized Visdom client.
        cam_crop (torch.Tensor): Cropped camera image tensor with shape (batch_size, C, H, W).
        warped_predict (torch.Tensor): Warped predicted image tensor with shape (batch_size, C, H, W).
        prj_GT (torch.Tensor): Ground Truth projected image tensor with shape (batch_size, C, H, W).
        batch_size (int): Batch size.
        step (int): Current training or validation step.
        phase (str, optional): Current phase, 'train' or 'valid'. Default is 'train'.
        max_nrow (int, optional): Maximum number of images per row. Default is 5.
        win (str, optional): Unique identifier for the Visdom window. If None, it will be auto-generated.
        save_path (str, optional): Path to save the images locally. If None, images will not be saved. Default is None.

    Returns:
        None
    """
    with torch.no_grad():
        # Select the first max_nrow samples based on batch_size
        if batch_size > max_nrow:
            # print(f"Batch size {batch_size} > {max_nrow}, selecting first {max_nrow} samples for visualization.")
            cam_crop = cam_crop[:max_nrow]
            warped_predict = warped_predict[:max_nrow]
            prj_GT = prj_GT[:max_nrow]
        else:
            print(f"Batch size {batch_size} <= {max_nrow}, using all {batch_size} samples for visualization.")

        # Concatenate the image tensors
        images = torch.cat([cam_crop, warped_predict, prj_GT], dim=0)
        # print(f"Concatenated image shape: {images.shape}")  # Expected shape (3 * selected_count, C, H, W)

        # Create the image grid
        nrow = min(batch_size, max_nrow)
        images_grid = vutils.make_grid(images, nrow=nrow, normalize=True, scale_each=True)
        # print(f"Image grid shape: {images_grid.shape}")  # Expected shape (C, H, W)

        # Generate window ID if not provided
        if win is None:
            win = f"{phase}_{step}"

        # Set the title
        full_title = f"{phase.capitalize()}_Step: {step}"

        # Display the images in Visdom
        vis.image(
            images_grid.cpu().numpy(),  # No need to swap dimensions
            opts=dict(title=full_title),
            win=win
        )
        # print(f"Displayed in Visdom window '{win}': {full_title}")

        # Save the images to local file if save_path is provided
        if save_path is not None:
            os.makedirs(os.path.dirname(save_path), exist_ok=True)  # Ensure the save directory exists
            vutils.save_image(images, save_path, nrow=nrow, normalize=True, scale_each=True)
            # print(f"Images saved to: {save_path}"


def log_test_metrics(logger, step, metrics):
    """
    Log test metrics to WandB.
    Parameters:
    - logger (WandBLogger): WandB logger instance
    - step (int): Current step, used as the x-coordinate
    - metrics (dict): Metrics dictionary to log, including 'test_mae', 'test_ssim', 'test_psnr', 'test_rmse', 'test_deltaE', 'test_lpips'
    """
    # Log each individual metric
    for key, value in metrics.items():
        logger.log_metrics(step, {key: value})

        # Calculate and log total loss
    total_loss = (
            metrics['test_mae'] +
            metrics['test_ssim'] +
            metrics['test_rmse'] +
            metrics['test_deltaE'] +
            metrics['test_lpips']
    )
    logger.log_metrics(step, {'test_loss': total_loss})
